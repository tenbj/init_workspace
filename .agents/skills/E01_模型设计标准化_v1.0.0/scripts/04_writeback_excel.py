from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from e01_common import (
    clean_text,
    copy_file,
    read_json,
    read_jsonl,
    rel,
    to_project_path,
    write_csv,
    write_json,
)


REVIEW_SHEETS = ["AI字段建议", "待人工确认", "待沉淀标准", "运行摘要"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Write E01 AI decisions back into review workbook and CSV deliverables.")
    parser.add_argument("--run-dir", required=True)
    parser.add_argument("--mode", default="review_only", choices=["review_only"])
    parser.add_argument("--allow-validation-errors", action="store_true")
    return parser.parse_args()


def write_rows(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    ws.freeze_panes = "A2"
    for idx, header in enumerate(headers, 1):
        width = max(12, min(48, len(header) * 2 + 4))
        ws.column_dimensions[ws.cell(1, idx).column_letter].width = width


def build_review_rows(fields: list[dict[str, Any]], decisions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    lookup = {(f["sheet_name"], int(f["row_number"])): f for f in fields}
    rows: list[dict[str, Any]] = []
    for decision in decisions:
        key = (clean_text(decision.get("sheet_name")), int(decision.get("row_number")))
        field = lookup.get(key, {})
        rows.append({
            "sheet名": key[0],
            "行号": key[1],
            "字段序号": field.get("field_order", ""),
            "字段分类": field.get("field_category", ""),
            "二级分类": field.get("field_subcategory", ""),
            "字段描述": field.get("field_desc", ""),
            "当前字段名": decision.get("current_field_name", field.get("field_name", "")),
            "推荐字段名": decision.get("recommended_field_name", ""),
            "数据类型": field.get("data_type", ""),
            "参考字段": field.get("reference_field", ""),
            "参考库表": field.get("reference_table", ""),
            "决策": decision.get("decision", ""),
            "置信度": decision.get("confidence", ""),
            "AI理由": decision.get("reason", ""),
            "人工问题": decision.get("human_question", ""),
            "是否允许回写": decision.get("writeback_allowed", False),
        })
    return rows


def build_confirm_rows(review_items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in review_items:
        rows.append({
            "问题类型": item.get("issue_type", ""),
            "严重级别": item.get("severity", ""),
            "sheet名": item.get("sheet_name", ""),
            "行号": item.get("row_number", ""),
            "问题描述": item.get("issue", ""),
            "AI建议": item.get("suggestion", ""),
            "AI理由": item.get("ai_reason", ""),
            "用户意见": "",
            "处理状态": "待确认",
        })
    return rows


def build_gap_rows(gaps: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for gap in gaps:
        rows.append({
            "缺口类型": gap.get("gap_type", ""),
            "来源sheet": gap.get("source_sheet", ""),
            "来源行号": gap.get("source_row", ""),
            "字段描述": gap.get("field_desc", ""),
            "推荐标准": gap.get("suggested_standard", ""),
            "AI理由": gap.get("reason", ""),
            "状态": gap.get("status", "待确认"),
            "用户意见": "",
        })
    return rows


def main() -> None:
    args = parse_args()
    run_dir = to_project_path(args.run_dir)
    manifest = read_json(run_dir / "run_manifest.json")
    validation_path = run_dir / "04_validation" / "validation_report.json"
    if validation_path.exists():
        validation = read_json(validation_path)
        if validation.get("status") == "failed" and not args.allow_validation_errors:
            raise RuntimeError("validation_report.json is failed; rerun with --allow-validation-errors to write anyway.")

    model_path = to_project_path(manifest["inputs"]["model_design"])
    fields = read_jsonl(run_dir / "01_parse" / "entity_fields.jsonl")
    decisions = read_jsonl(run_dir / "03_ai_decisions" / "ai_field_review_decisions.jsonl")
    review_items = read_jsonl(run_dir / "04_validation" / "review_items.jsonl")
    gaps = read_jsonl(run_dir / "03_ai_decisions" / "ai_standard_gaps.jsonl")

    if not decisions:
        raise RuntimeError("No AI field decisions found. Fill 03_ai_decisions/ai_field_review_decisions.jsonl first.")

    review_rows = build_review_rows(fields, decisions)
    confirm_rows = build_confirm_rows(review_items)
    gap_rows = build_gap_rows(gaps)

    draft_path = run_dir / "05_writeback" / "workbook_draft.xlsx"
    final_path = run_dir / "06_deliverables" / "模型设计标准化建议版.xlsx"
    copy_file(model_path, draft_path)
    copy_file(model_path, final_path)

    wb = load_workbook(final_path)
    for sheet_name in REVIEW_SHEETS:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    write_rows(
        wb.create_sheet("AI字段建议"),
        ["sheet名", "行号", "字段序号", "字段分类", "二级分类", "字段描述", "当前字段名", "推荐字段名", "数据类型", "参考字段", "参考库表", "决策", "置信度", "AI理由", "人工问题", "是否允许回写"],
        review_rows,
    )
    write_rows(
        wb.create_sheet("待人工确认"),
        ["问题类型", "严重级别", "sheet名", "行号", "问题描述", "AI建议", "AI理由", "用户意见", "处理状态"],
        confirm_rows,
    )
    write_rows(
        wb.create_sheet("待沉淀标准"),
        ["缺口类型", "来源sheet", "来源行号", "字段描述", "推荐标准", "AI理由", "状态", "用户意见"],
        gap_rows,
    )
    summary_rows = [
        {"指标": "运行目录", "值": rel(run_dir)},
        {"指标": "模型设计", "值": manifest["inputs"]["model_design"]},
        {"指标": "字段决策数", "值": len(decisions)},
        {"指标": "待人工确认数", "值": len(confirm_rows)},
        {"指标": "待沉淀标准数", "值": len(gap_rows)},
        {"指标": "输出模式", "值": args.mode},
        {"指标": "生成时间", "值": datetime.now().isoformat(timespec="seconds")},
    ]
    write_rows(wb.create_sheet("运行摘要"), ["指标", "值"], summary_rows)
    wb.save(final_path)

    write_csv(
        run_dir / "06_deliverables" / "待人工确认清单.csv",
        confirm_rows,
        ["问题类型", "严重级别", "sheet名", "行号", "问题描述", "AI建议", "AI理由", "用户意见", "处理状态"],
    )
    write_csv(
        run_dir / "06_deliverables" / "待沉淀标准清单.csv",
        gap_rows,
        ["缺口类型", "来源sheet", "来源行号", "字段描述", "推荐标准", "AI理由", "状态", "用户意见"],
    )

    plan = {
        "mode": args.mode,
        "actions": [
            {"action": "copy_source_workbook", "target": rel(final_path)},
            {"action": "replace_review_sheets", "sheets": REVIEW_SHEETS},
            {"action": "export_csv", "files": ["待人工确认清单.csv", "待沉淀标准清单.csv"]},
        ],
        "overwrite_original_field_names": False,
    }
    log = {
        "status": "completed",
        "review_rows": len(review_rows),
        "confirm_rows": len(confirm_rows),
        "gap_rows": len(gap_rows),
        "draft_workbook": rel(draft_path),
        "final_workbook": rel(final_path),
    }
    write_json(run_dir / "05_writeback" / "writeback_plan.json", plan)
    write_json(run_dir / "05_writeback" / "writeback_log.json", log)

    manifest["status"] = "writeback_completed"
    manifest["outputs"]["review_excel"] = rel(final_path)
    manifest["outputs"]["confirm_csv"] = rel(run_dir / "06_deliverables" / "待人工确认清单.csv")
    manifest["outputs"]["standard_gap_csv"] = rel(run_dir / "06_deliverables" / "待沉淀标准清单.csv")
    write_json(run_dir / "run_manifest.json", manifest)

    print(f"[OK] review_excel={rel(final_path)}")
    print(f"[OK] confirm_rows={len(confirm_rows)} standard_gap_rows={len(gap_rows)}")


if __name__ == "__main__":
    main()
