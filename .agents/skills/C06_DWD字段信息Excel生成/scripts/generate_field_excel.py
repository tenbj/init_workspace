"""
DWD 字段信息 Excel 生成脚本

用法：
    python generate_field_excel.py \\
        --dwd-sql  <dwd_insert_sql_path> \\
        --ods-sql  <ods_create_sql_path> \\
        --dim-sql  <dim1_create_sql_path> \\
        --dim-sql  <dim2_create_sql_path> \\
        --output   <output_xlsx_path>

说明：
    - --dwd-sql  DWD 的 INSERT OVERWRITE SQL 文件（必填）
    - --ods-sql  ODS 表的 CREATE TABLE SQL 文件（必填）
    - --dim-sql  DIM 表的 CREATE TABLE SQL 文件（可重复，0 个或多个）
    - --output   输出 xlsx 文件路径（必填）
"""

import sys
import io
import re
import argparse
import pathlib

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] 缺少依赖：pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ── 常量 ──────────────────────────────────────────────────────────────────────
SYSTEM_FIELDS   = {"first_inserted_dt", "etl_dt", "last_updated_dt"}
SYSTEM_ORDER    = ["first_inserted_dt", "etl_dt", "last_updated_dt"]

HEADERS     = ["排序序号", "字段类型", "Type", "字段原名", "字段名",
               "字段释义", "字段分组", "Nullable", "物理主键", "业务主键", "字段备注"]
COL_WIDTHS  = [10, 10, 20, 28, 20, 36, 12, 10, 10, 10, 50]

# ── SQL 解析工具 ───────────────────────────────────────────────────────────────

def sql_type_to_category(sql_type: str) -> str:
    t = sql_type.lower()
    if t.startswith("datetime") or t in ("date", "timestamp"):
        return "日期"
    if (t in ("bigint", "int", "tinyint", "smallint", "mediumint")
            or t.startswith("decimal") or t.startswith("float") or t.startswith("double")):
        return "数值"
    return "文本"


def split_comment(comment: str):
    """按第一个逗号分割 COMMENT，返回 (中文名, 备注/示例)。"""
    if not comment:
        return "", None
    idx = comment.find(",")
    if idx == -1:
        return comment.strip(), None
    return comment[:idx].strip(), comment[idx + 1:].strip()


def parse_create_table(sql_text: str) -> dict:
    """
    解析 CREATE TABLE SQL，返回：
    {
        "table_name": str,
        "pk_field":   str,
        "fields":     {field_name: {"sql_type", "nullable", "comment"}}
    }
    """
    # 表名
    tbl_match = re.search(r"CREATE TABLE `(\w+)`", sql_text)
    table_name = tbl_match.group(1) if tbl_match else ""

    # 主键
    pk_match = re.search(r"UNIQUE KEY\(`(\w+)`\)", sql_text)
    pk_field = pk_match.group(1) if pk_match else ""

    # 字段
    field_re = re.compile(
        r"`(\w+)`\s+([\w()]+(?:,\d+)?(?:,\d+)?)\s+(NOT NULL|NULL)"
        r'(?:\s+COMMENT\s+"((?:[^"\\]|\\.)*)")?',
        re.MULTILINE,
    )
    fields = {}
    for name, sql_type, nullability, comment in field_re.findall(sql_text):
        fields[name] = {
            "sql_type":  sql_type,
            "nullable":  "否" if nullability == "NOT NULL" else "是",
            "comment":   comment or "",
        }

    return {"table_name": table_name, "pk_field": pk_field, "fields": fields}


def parse_dwd_sql(sql_text: str):
    """
    解析 DWD INSERT SQL，返回：
    {
        "insert_fields": [field_name, ...],   # INSERT 括号内的字段，原始顺序
        "select_map":    {dwd_field: expr},   # dwd_field → 原始 SELECT 表达式
        "aliases":       {alias: schema.table} # FROM/JOIN 中的表别名
    }
    """
    # ── INSERT 字段列表 ──
    ins_match = re.search(r"insert\s+overwrite\s+table\s+\S+\s*\((.*?)\)\s*select",
                          sql_text, re.IGNORECASE | re.DOTALL)
    if not ins_match:
        print("[ERROR] 未找到 INSERT 字段列表", file=sys.stderr)
        sys.exit(1)
    insert_fields = [f.strip() for f in ins_match.group(1).split(",") if f.strip()]

    # ── SELECT 表达式 → alias 映射 ──
    # 找 SELECT ... FROM 之间的部分
    sel_match = re.search(r"\bselect\b(.*?)\bfrom\b", sql_text, re.IGNORECASE | re.DOTALL)
    select_map = {}
    if sel_match:
        sel_body = sel_match.group(1)
        # 每个 SELECT 项以逗号分隔（忽略括号内逗号）
        items = _split_select_items(sel_body)
        for item in items:
            item = item.strip().rstrip(",").strip()
            # 提取 alias：末尾 `as alias` 或裸 alias
            alias_match = re.search(r"\bas\s+(\w+)\s*$", item, re.IGNORECASE)
            if alias_match:
                alias = alias_match.group(1)
                expr  = item[: alias_match.start()].strip()
            else:
                # 无 as，最后一个词可能是 alias，或者 src.field 中 field 本身
                parts = item.rsplit(".", 1)
                alias = parts[-1].strip()
                expr  = item
            select_map[alias] = expr

    # ── FROM / JOIN 表别名 ──
    aliases = {}
    # FROM schema.table alias 或 FROM schema.table AS alias
    from_re = re.compile(
        r"\b(?:from|join)\s+(\w+\.\w+|\w+)\s+(?:as\s+)?(\w+)\b",
        re.IGNORECASE,
    )
    for schema_table, alias in from_re.findall(sql_text):
        aliases[alias] = schema_table

    return {"insert_fields": insert_fields, "select_map": select_map, "aliases": aliases}


def _split_select_items(sel_body: str) -> list:
    """按顶层逗号分割 SELECT 表达式列表（忽略括号内的逗号）。"""
    items, depth, buf = [], 0, []
    for ch in sel_body:
        if ch in "([":
            depth += 1
            buf.append(ch)
        elif ch in ")]":
            depth -= 1
            buf.append(ch)
        elif ch == "," and depth == 0:
            items.append("".join(buf).strip())
            buf = []
        else:
            buf.append(ch)
    if buf:
        items.append("".join(buf).strip())
    return items


# ── 字段行构建 ────────────────────────────────────────────────────────────────

def build_field_row(dwd_field: str, expr: str, ods: dict, dim_tables: list, seq: int) -> dict:
    """
    根据 SELECT 表达式和可用的 CREATE TABLE 信息，构建一行字段数据。
    """
    expr_lower = expr.lower().strip()

    # ── 1. NOW() → 系统 datetime ──
    if re.match(r"^now\(\)", expr_lower):
        return _make_row(seq, dwd_field, "datetime(3)", "日期", dwd_field, dwd_field, "是", "否", None)

    # ── 2. NULL as field → 新增占位字段 ──
    if re.match(r"^null\s*$", expr_lower.split("as")[0].strip() if " as " in expr_lower else expr_lower):
        return _make_row(seq, dwd_field, "varchar(255)", "文本", _cn_name(dwd_field), _cn_name(dwd_field),
                         "是", "否", "新增字段，待补充")

    # ── 3. from_unixtime → datetime 转换 ──
    if "from_unixtime" in expr_lower:
        src_field = _extract_src_field(expr)
        note = f"由 ODS {src_field} (unix毫秒戳) 转换" if src_field else "由 ODS unix毫秒戳转换"
        return _make_row(seq, dwd_field, "datetime(3)", "日期", _cn_name(dwd_field), _cn_name(dwd_field),
                         "是", "否", note)

    # ── 4. date(...) → date 类型 ──
    if re.match(r"^date\s*\(", expr_lower):
        return _make_row(seq, dwd_field, "date", "日期", _cn_name(dwd_field), _cn_name(dwd_field), "是", "否", None)

    # ── 5. src.field 或 dim_alias.field → 查 ODS/DIM ──
    dot_match = re.match(r"^(\w+)\.(\w+)\s*$", expr.strip())
    if dot_match:
        alias, src_field = dot_match.group(1), dot_match.group(2)

        # 先查 ODS
        if src_field in ods["fields"]:
            f      = ods["fields"][src_field]
            cn, note = split_comment(f["comment"])
            is_pk  = "是" if src_field == ods["pk_field"] else "否"
            nullable = f["nullable"]
            # 对于系统字段，nullable 固定为 是（即使 ODS 写了 NOT NULL）
            if dwd_field in SYSTEM_FIELDS:
                nullable = "是"
                is_pk    = "否"
            return _make_row(seq, dwd_field, f["sql_type"],
                             sql_type_to_category(f["sql_type"]),
                             cn or dwd_field, cn or dwd_field,
                             nullable, is_pk, note)

        # 再查 DIM 表
        for dim in dim_tables:
            if src_field in dim["fields"]:
                f        = dim["fields"][src_field]
                cn, note = split_comment(f["comment"])
                return _make_row(seq, dwd_field, f["sql_type"],
                                 sql_type_to_category(f["sql_type"]),
                                 cn or dwd_field, cn or dwd_field,
                                 f["nullable"], "否", note)

    # ── 6. 兜底：按字段名推断 ──
    inferred = _infer_by_name(dwd_field)
    return _make_row(seq, dwd_field, inferred["sql_type"], inferred["category"],
                     _cn_name(dwd_field), _cn_name(dwd_field), "是", "否", None)


def _make_row(seq, orig, sql_type, category, name, desc, nullable, phys_pk, note):
    return {
        "seq": seq, "category": category, "sql_type": sql_type,
        "orig": orig, "name": name, "desc": desc,
        "group": None, "nullable": nullable,
        "phys_pk": phys_pk, "biz_pk": "否", "note": note,
    }


def _cn_name(field_name: str) -> str:
    """无 COMMENT 时用字段名本身作为中文名。"""
    return field_name


def _extract_src_field(expr: str) -> str:
    """从 from_unixtime(cast(src.field as bigint)/1000) 提取 src_field。"""
    m = re.search(r"(\w+)\.(\w+)", expr)
    return m.group(2) if m else ""


def _infer_by_name(name: str) -> dict:
    """按字段名后缀推断类型（兜底）。"""
    n = name.lower()
    if n.endswith("_dt") or n.endswith("_time") or n.endswith("_date") or "date" in n:
        return {"sql_type": "datetime(3)", "category": "日期"}
    if n.endswith("_id") or n.endswith("_cnt") or n.endswith("_qty"):
        return {"sql_type": "bigint", "category": "数值"}
    if any(kw in n for kw in ("amount", "cost", "price", "rate", "ratio")):
        return {"sql_type": "decimal(18,6)", "category": "数值"}
    return {"sql_type": "text", "category": "文本"}


# ── Excel 写出 ────────────────────────────────────────────────────────────────

def write_excel(rows: list, output_path: pathlib.Path, sheet_name: str = "字段信息"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel 限制 sheet 名 ≤31 字符

    h_font  = Font(bold=True, name="微软雅黑", size=10)
    h_fill  = PatternFill(patternType="solid", fgColor="BDD7EE")
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border  = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))
    d_font   = Font(name="微软雅黑", size=10)
    d_center = Alignment(horizontal="center", vertical="center")
    d_left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    fill_w   = PatternFill(patternType="solid", fgColor="FFFFFF")
    fill_g   = PatternFill(patternType="solid", fgColor="F2F2F2")

    ws.row_dimensions[1].height = 22
    for ci, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = h_font, h_fill, h_align, border
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, row in enumerate(rows, 2):
        fill = fill_w if ri % 2 == 0 else fill_g
        ws.row_dimensions[ri].height = 18
        vals = [row["seq"], row["category"], row["sql_type"], row["orig"],
                row["name"], row["desc"], row["group"], row["nullable"],
                row["phys_pk"], row["biz_pk"], row["note"]]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font, cell.fill, cell.border = d_font, fill, border
            cell.alignment = d_left if ci in (5, 6, 11) else d_center

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ── 主流程 ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="生成 DWD 字段信息 Excel")
    parser.add_argument("--dwd-sql",  required=True, help="DWD INSERT SQL 文件路径")
    parser.add_argument("--ods-sql",  required=True, help="ODS CREATE TABLE SQL 文件路径")
    parser.add_argument("--dim-sql",  action="append", default=[], help="DIM CREATE TABLE SQL 文件路径（可重复）")
    parser.add_argument("--output",   required=True, help="输出 xlsx 文件路径")
    args = parser.parse_args()

    dwd_text = pathlib.Path(args.dwd_sql).read_text(encoding="utf-8")
    ods_text = pathlib.Path(args.ods_sql).read_text(encoding="utf-8")
    dim_texts = [pathlib.Path(p).read_text(encoding="utf-8") for p in args.dim_sql]

    ods = parse_create_table(ods_text)
    dim_tables = [parse_create_table(t) for t in dim_texts]
    dwd = parse_dwd_sql(dwd_text)

    print(f"[INFO] ODS 表：{ods['table_name']}，字段数：{len(ods['fields'])}")
    for d in dim_tables:
        print(f"[INFO] DIM 表：{d['table_name']}，字段数：{len(d['fields'])}")
    print(f"[INFO] DWD INSERT 字段数：{len(dwd['insert_fields'])}")

    sys_rows, biz_rows = [], []
    biz_seq = 1

    for field in dwd["insert_fields"]:
        expr = dwd["select_map"].get(field, field)
        if field in SYSTEM_FIELDS:
            row = build_field_row(field, expr, ods, dim_tables, 0)
            sys_rows.append(row)
        else:
            row = build_field_row(field, expr, ods, dim_tables, biz_seq)
            biz_rows.append(row)
            biz_seq += 1

    # 系统字段按固定顺序排列
    sys_rows.sort(key=lambda r: SYSTEM_ORDER.index(r["orig"])
                  if r["orig"] in SYSTEM_ORDER else 99)

    all_rows = sys_rows + biz_rows
    output_path = pathlib.Path(args.output)
    write_excel(all_rows, output_path, sheet_name=output_path.stem)

    print(f"[OK] 已生成：{output_path}")
    print(f"     总字段：{len(all_rows)}  系统字段：{len(sys_rows)}  业务字段：{len(biz_rows)}")


if __name__ == "__main__":
    main()
