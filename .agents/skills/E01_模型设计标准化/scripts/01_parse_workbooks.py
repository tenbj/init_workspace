from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from e01_common import (
    FIELD_HEADERS,
    clean_text,
    create_run_dir,
    declared_table_from_first_row,
    ensure_run_subdirs,
    file_manifest,
    find_header_row,
    get_by_header,
    header_map,
    load_xlsx_xml,
    markdown_table,
    rel,
    row_values,
    rows_to_records,
    safe_task_name,
    source_cell,
    to_project_path,
    write_json,
    write_jsonl,
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Parse standard library and model design workbooks into E01 run artifacts."
    )
    parser.add_argument("--standard-library", default="input/标准库.xlsx")
    parser.add_argument("--model-design", required=True)
    parser.add_argument("--sample-design")
    parser.add_argument("--data-assets")
    parser.add_argument("--run-dir")
    parser.add_argument("--output-root")
    parser.add_argument("--task-name")
    parser.add_argument("--include-hidden-entity-sheets", action="store_true")
    return parser.parse_args()


def workbook_summary(workbook: dict[str, Any]) -> dict[str, Any]:
    return {
        "file_path": workbook["file_path"],
        "file_hash": workbook["file_hash"],
        "sheet_count": workbook["sheet_count"],
        "sheets": [
            {
                "sheet_id": sheet["sheet_id"],
                "sheet_name": sheet["sheet_name"],
                "visible": sheet["visible"],
                "state": sheet["state"],
                "non_empty_rows": sheet["non_empty_rows"],
                "max_row_seen": sheet["max_row_seen"],
                "max_col_seen": sheet["max_col_seen"],
                "detected_type": sheet["detected_type"],
            }
            for sheet in workbook["sheets"]
        ],
    }


def inspect_workbook_layout(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=False, read_only=False)
    sheets: list[dict[str, Any]] = []
    for index, ws in enumerate(wb.worksheets, 1):
        hidden_rows = [idx for idx, dim in ws.row_dimensions.items() if dim.hidden]
        hidden_columns = [key for key, dim in ws.column_dimensions.items() if dim.hidden]
        column_dimensions = {
            key: {
                "width": dim.width,
                "hidden": bool(dim.hidden),
                "outline_level": dim.outlineLevel,
            }
            for key, dim in ws.column_dimensions.items()
            if dim.width or dim.hidden or dim.outlineLevel
        }
        row_dimensions = {
            str(key): {
                "height": dim.height,
                "hidden": bool(dim.hidden),
                "outline_level": dim.outlineLevel,
            }
            for key, dim in ws.row_dimensions.items()
            if dim.height or dim.hidden or dim.outlineLevel
        }
        non_empty_style_ids: dict[str, int] = {}
        font_samples: dict[str, dict[str, Any]] = {}
        fill_samples: dict[str, dict[str, Any]] = {}
        number_formats: dict[str, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                style_id = str(cell.style_id)
                non_empty_style_ids[style_id] = non_empty_style_ids.get(style_id, 0) + 1
                if style_id not in font_samples:
                    font_samples[style_id] = {
                        "name": cell.font.name,
                        "size": cell.font.sz,
                        "bold": bool(cell.font.bold),
                        "italic": bool(cell.font.italic),
                        "color": cell.font.color.rgb if cell.font.color and cell.font.color.type == "rgb" else None,
                    }
                fill_id = str(cell.fill.fill_type or "none")
                if fill_id not in fill_samples:
                    fill_samples[fill_id] = {
                        "fill_type": cell.fill.fill_type,
                        "fgColor": cell.fill.fgColor.rgb if cell.fill.fgColor and cell.fill.fgColor.type == "rgb" else None,
                    }
                fmt = cell.number_format or ""
                number_formats[fmt] = number_formats.get(fmt, 0) + 1
        sheets.append({
            "sheet_index": index,
            "sheet_name": ws.title,
            "state": ws.sheet_state,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else "",
            "auto_filter_ref": ws.auto_filter.ref,
            "merged_cells": [str(rng) for rng in ws.merged_cells.ranges],
            "hidden_rows": hidden_rows,
            "hidden_columns": hidden_columns,
            "column_dimensions": column_dimensions,
            "row_dimensions": row_dimensions,
            "non_empty_style_id_counts": non_empty_style_ids,
            "font_samples_by_style_id": font_samples,
            "fill_samples": fill_samples,
            "number_format_counts": number_formats,
        })
    return {
        "file_path": rel(path),
        "sheet_count": len(wb.worksheets),
        "active_sheet": wb.active.title if wb.active else "",
        "defined_names": [name for name in wb.defined_names],
        "sheets": sheets,
    }


def parse_model_catalog(model_wb: dict[str, Any]) -> list[dict[str, Any]]:
    catalog_sheets = [s for s in model_wb["sheets"] if s["detected_type"] == "model_catalog"]
    if not catalog_sheets:
        return []
    sheet = catalog_sheets[0]
    header = find_header_row(sheet["rows"], ["模型编号", "模型名称", "实体表名 ( 库名.表名)"])
    if not header:
        header = find_header_row(sheet["rows"], ["模型编号", "模型名称"])
    if not header:
        return []
    hmap = header_map(header)
    records: list[dict[str, Any]] = []
    for row in sheet["rows"]:
        if row["row_number"] <= header["row_number"]:
            continue
        record = {
            "row_number": row["row_number"],
            "模型编号": get_by_header(row, hmap, "模型编号"),
            "所属数据域": get_by_header(row, hmap, "所属数据域"),
            "专项": get_by_header(row, hmap, "专项"),
            "数仓分层": get_by_header(row, hmap, "数仓分层"),
            "模型名称": get_by_header(row, hmap, "模型名称"),
            "当前状态": get_by_header(row, hmap, "当前状态"),
            "实体表名": (
                get_by_header(row, hmap, "实体表名(库名.表名)")
                or get_by_header(row, hmap, "实体表名 ( 库名.表名)")
                or get_by_header(row, hmap, "实体表名")
            ),
            "备注": get_by_header(row, hmap, "备注"),
        }
        if any(v for k, v in record.items() if k != "row_number"):
            records.append(record)
    return records


def parse_dependency(model_wb: dict[str, Any]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for sheet in model_wb["sheets"]:
        if sheet["detected_type"] != "dependency":
            continue
        header = find_header_row(sheet["rows"], ["数据来源", "模型名称"])
        if not header and sheet["rows"]:
            header = sheet["rows"][0]
        if not header:
            continue
        max_col = sheet["max_col_seen"]
        headers = row_values(header, max_col)
        for row in sheet["rows"]:
            if row["row_number"] <= header["row_number"]:
                continue
            values = row_values(row, max_col)
            record = {
                headers[i] if i < len(headers) and headers[i] else f"col_{i + 1}": values[i]
                for i in range(max_col)
                if i < len(values)
            }
            if any(record.values()):
                record["_sheet_name"] = sheet["sheet_name"]
                record["_row_number"] = row["row_number"]
                result.append(record)
    return result


def parse_standard_library(std_wb: dict[str, Any]) -> dict[str, Any]:
    result = {"rules": [], "mappings": [], "other": []}
    for sheet in std_wb["sheets"]:
        records = rows_to_records(sheet["rows"], "序号")
        item = {
            "sheet_name": sheet["sheet_name"],
            "detected_type": sheet["detected_type"],
            "rows": records,
        }
        if sheet["detected_type"] == "standard_rules":
            result["rules"].extend(records)
        elif sheet["detected_type"] == "standard_mapping":
            result["mappings"].extend(records)
        else:
            result["other"].append(item)
    return result


def should_parse_entity_sheet(sheet: dict[str, Any], include_hidden: bool) -> bool:
    if sheet["detected_type"] != "entity_design":
        return False
    if not include_hidden and not sheet["visible"]:
        return False
    if "模板" in sheet["sheet_name"] or sheet["sheet_name"].startswith("附."):
        return False
    return True


def parse_entity_fields(model_wb: dict[str, Any], include_hidden: bool) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    entity_sheets: list[dict[str, Any]] = []
    fields: list[dict[str, Any]] = []

    for sheet in model_wb["sheets"]:
        if not should_parse_entity_sheet(sheet, include_hidden):
            continue
        header = find_header_row(sheet["rows"], ["字段名", "字段描述"])
        if not header:
            continue
        hmap = header_map(header)
        declared_table = declared_table_from_first_row(sheet)
        data_rows = [r for r in sheet["rows"] if r["row_number"] > header["row_number"]]
        parsed_count = 0

        for row in data_rows:
            record = {
                "sheet_name": sheet["sheet_name"],
                "row_number": row["row_number"],
                "declared_table": declared_table,
                "field_category": get_by_header(row, hmap, FIELD_HEADERS["field_category"]),
                "field_subcategory": get_by_header(row, hmap, FIELD_HEADERS["field_subcategory"]),
                "field_order": get_by_header(row, hmap, FIELD_HEADERS["field_order"]),
                "field_name": get_by_header(row, hmap, FIELD_HEADERS["field_name"]),
                "data_type": get_by_header(row, hmap, FIELD_HEADERS["data_type"]),
                "field_desc": get_by_header(row, hmap, FIELD_HEADERS["field_desc"]),
                "reference_field": get_by_header(row, hmap, FIELD_HEADERS["reference_field"]),
                "reference_table": get_by_header(row, hmap, FIELD_HEADERS["reference_table"]),
                "is_retained": get_by_header(row, hmap, FIELD_HEADERS["is_retained"]),
                "is_deleted": get_by_header(row, hmap, FIELD_HEADERS["is_deleted"]),
                "is_enum_field": get_by_header(row, hmap, FIELD_HEADERS["is_enum_field"]),
                "is_primary_key": get_by_header(row, hmap, FIELD_HEADERS["is_primary_key"]),
                "is_business_key": get_by_header(row, hmap, FIELD_HEADERS["is_business_key"]),
                "remark": get_by_header(row, hmap, FIELD_HEADERS["remark"]),
            }
            if not any(record[k] for k in [
                "field_order",
                "field_name",
                "field_desc",
                "reference_field",
                "reference_table",
                "remark",
            ]):
                continue
            record["is_continuation_row"] = (
                not record["field_name"]
                and not record["field_desc"]
                and bool(record["reference_field"] or record["reference_table"])
            )
            record["source_cell_map"] = {
                "field_name": source_cell(hmap, FIELD_HEADERS["field_name"], row["row_number"]),
                "field_desc": source_cell(hmap, FIELD_HEADERS["field_desc"], row["row_number"]),
                "reference_field": source_cell(hmap, FIELD_HEADERS["reference_field"], row["row_number"]),
                "reference_table": source_cell(hmap, FIELD_HEADERS["reference_table"], row["row_number"]),
                "is_retained": source_cell(hmap, FIELD_HEADERS["is_retained"], row["row_number"]),
                "is_deleted": source_cell(hmap, FIELD_HEADERS["is_deleted"], row["row_number"]),
                "is_enum_field": source_cell(hmap, FIELD_HEADERS["is_enum_field"], row["row_number"]),
            }
            fields.append(record)
            parsed_count += 1

        entity_sheets.append({
            "sheet_id": sheet["sheet_id"],
            "sheet_name": sheet["sheet_name"],
            "visible": sheet["visible"],
            "state": sheet["state"],
            "declared_table": declared_table,
            "header_row": header["row_number"],
            "data_start_row": min((r["row_number"] for r in data_rows), default=None),
            "data_end_row": max((r["row_number"] for r in data_rows), default=None),
            "field_row_count": parsed_count,
            "max_col_seen": sheet["max_col_seen"],
        })

    return entity_sheets, fields


def build_template_check(
    model_wb: dict[str, Any],
    std_wb: dict[str, Any],
    entity_sheets: list[dict[str, Any]],
) -> dict[str, Any]:
    issues: list[dict[str, Any]] = []
    model_types = {s["detected_type"] for s in model_wb["sheets"]}
    std_types = {s["detected_type"] for s in std_wb["sheets"]}

    if "model_catalog" not in model_types:
        issues.append({"severity": "error", "issue": "缺少模型设计目录 sheet"})
    if "dependency" not in model_types:
        issues.append({"severity": "warning", "issue": "缺少模型依赖说明/数据流 sheet"})
    if not entity_sheets:
        issues.append({"severity": "error", "issue": "没有识别到可见实体设计 sheet"})
    if "standard_rules" not in std_types:
        issues.append({"severity": "error", "issue": "标准库缺少规则标准库 sheet"})
    if "standard_mapping" not in std_types:
        issues.append({"severity": "error", "issue": "标准库缺少映射标准库 sheet"})

    required = ["字段名", "字段描述", "字段数据类型", "参考字段", "参考库表"]
    for sheet in model_wb["sheets"]:
        if sheet["detected_type"] != "entity_design" or not sheet["visible"]:
            continue
        header = find_header_row(sheet["rows"], ["字段名", "字段描述"])
        if not header:
            continue
        hmap = header_map(header)
        missing = [h for h in required if h.replace(" ", "") not in hmap]
        if missing:
            issues.append({
                "severity": "error",
                "sheet_name": sheet["sheet_name"],
                "issue": "实体设计 sheet 缺少必需列",
                "missing_headers": missing,
            })

    return {
        "status": "failed" if any(i["severity"] == "error" for i in issues) else "passed",
        "issue_count": len(issues),
        "issues": issues,
    }


def write_source_files(run_dir: Path, manifests: list[dict[str, Any]]) -> None:
    rows = [[m["role"], m["path"], m["size_bytes"], m["sha256"]] for m in manifests]
    content = "# 输入文件\n\n" + markdown_table(["类型", "路径", "大小", "sha256"], rows)
    (run_dir / "00_inputs" / "source_files.md").write_text(content, encoding="utf-8")


def main() -> None:
    args = parse_args()
    task_name = args.task_name or safe_task_name(args.model_design)
    run_dir = create_run_dir(args.run_dir, args.output_root, task_name)
    ensure_run_subdirs(run_dir)

    model_path = to_project_path(args.model_design)
    standard_path = to_project_path(args.standard_library)
    if not model_path.exists():
        raise FileNotFoundError(model_path)
    if not standard_path.exists():
        raise FileNotFoundError(standard_path)

    input_manifests = [
        file_manifest(args.standard_library, "standard_library"),
        file_manifest(args.model_design, "model_design"),
    ]
    if args.sample_design:
        input_manifests.append(file_manifest(args.sample_design, "sample_design"))
    if args.data_assets:
        data_assets_path = to_project_path(args.data_assets)
        input_manifests.append({
            "role": "data_assets",
            "path": rel(data_assets_path),
            "exists": data_assets_path.exists(),
            "size_bytes": None,
            "sha256": None,
        })

    write_json(run_dir / "00_inputs" / "input_manifest.json", {
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "inputs": input_manifests,
    })
    write_source_files(run_dir, input_manifests)

    model_wb = load_xlsx_xml(model_path, "model_design")
    std_wb = load_xlsx_xml(standard_path, "standard_library")

    model_catalog = parse_model_catalog(model_wb)
    dependency_raw = parse_dependency(model_wb)
    standard_library_raw = parse_standard_library(std_wb)
    entity_sheets, entity_fields = parse_entity_fields(model_wb, args.include_hidden_entity_sheets)
    template_check = build_template_check(model_wb, std_wb, entity_sheets)

    workbook_manifest = {
        "model_design": workbook_summary(model_wb),
        "standard_library": workbook_summary(std_wb),
    }
    workbook_layout = {
        "model_design": inspect_workbook_layout(model_path),
        "standard_library": inspect_workbook_layout(standard_path),
    }

    write_json(run_dir / "01_parse" / "workbook_manifest.json", workbook_manifest)
    write_json(run_dir / "01_parse" / "workbook_layout.json", workbook_layout)
    write_json(run_dir / "01_parse" / "template_check.json", template_check)
    write_json(run_dir / "01_parse" / "model_catalog.json", model_catalog)
    write_json(run_dir / "01_parse" / "dependency_raw.json", dependency_raw)
    write_json(run_dir / "01_parse" / "standard_library_raw.json", standard_library_raw)
    write_jsonl(run_dir / "01_parse" / "entity_sheets.jsonl", entity_sheets)
    write_jsonl(run_dir / "01_parse" / "entity_fields.jsonl", entity_fields)

    run_manifest = {
        "run_id": run_dir.name,
        "skill": "E01_模型设计标准化",
        "mode": "review_only",
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "run_dir": rel(run_dir),
        "inputs": {
            "standard_library": rel(standard_path),
            "model_design": rel(model_path),
            "sample_design": rel(to_project_path(args.sample_design)) if args.sample_design else None,
            "data_assets": rel(to_project_path(args.data_assets)) if args.data_assets else None,
        },
        "outputs": {
            "review_excel": "06_deliverables/模型设计标准化建议版.xlsx",
            "summary": "run_summary.md",
        },
        "status": "parsed",
        "counts": {
            "entity_sheet_count": len(entity_sheets),
            "entity_field_row_count": len(entity_fields),
            "standard_rule_count": len(standard_library_raw["rules"]),
            "standard_mapping_count": len(standard_library_raw["mappings"]),
            "template_issue_count": template_check["issue_count"],
        },
    }
    write_json(run_dir / "run_manifest.json", run_manifest)

    print(f"[OK] run_dir={rel(run_dir)}")
    print(f"[OK] entity_sheets={len(entity_sheets)} entity_field_rows={len(entity_fields)}")
    print(f"[OK] template_check={template_check['status']} issues={template_check['issue_count']}")


if __name__ == "__main__":
    main()
