from __future__ import annotations

import argparse
import re
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

from e01_common import (
    clean_text,
    copy_file,
    read_json,
    read_jsonl,
    rel,
    sha256_file,
    to_project_path,
    write_csv,
    write_json,
)


REVIEW_SHEETS = ["字段标准化建议", "表名标准化建议", "待人工确认", "待沉淀标准", "运行摘要"]
LEGACY_REVIEW_SHEETS = ["AI字段建议", "AI表名建议"]
AUDIT_SHEET = "0 标准化结果"
AUTO_FIELD_DECISIONS = {"keep", "fill_safe", "replace_suggestion"}
AUTO_TABLE_DECISIONS = {"keep", "fill_safe", "replace_suggestion"}
CHANGED_FILL = PatternFill("solid", fgColor="FFF2CC")
KEPT_FILL = PatternFill("solid", fgColor="E2F0D9")
PENDING_FILL = PatternFill("solid", fgColor="FCE4D6")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HYPERLINK_FONT = Font(color="0563C1", underline="single")

DECISION_LABELS = {
    "keep": "保留",
    "fill_safe": "安全补充",
    "suggest_with_review": "建议人工复核",
    "replace_suggestion": "建议替换",
    "need_human_confirm": "需要人工确认",
    "cannot_decide": "无法判断",
}

CONFIDENCE_LABELS = {
    "high": "高",
    "medium": "中",
    "low": "低",
}

SEVERITY_LABELS = {
    "error": "错误",
    "warning": "警告",
}

ISSUE_TYPE_LABELS = {
    **DECISION_LABELS,
    "missing_field_decisions": "缺少字段决策文件",
    "missing_keys": "缺少必填字段",
    "invalid_row_number": "行号不是整数",
    "unexpected_field_row": "指向未解析字段行",
    "invalid_decision": "非法决策",
    "invalid_confidence": "非法置信度",
    "missing_reason": "缺少说明",
    "invalid_writeback_allowed": "非法回写标记",
    "field_name_format": "字段名格式不规范",
    "unsafe_writeback": "不安全回写",
    "missing_field_row_decision": "缺少字段行决策",
    "duplicate_recommended_field": "推荐字段名重复",
    "missing_gap_key": "待沉淀标准缺少字段",
}

GAP_TYPE_LABELS = {
    "new_word_root": "新词根",
    "semantic_relation": "语义关系",
    "alias": "别名",
    "rule_gap": "规则缺口",
    "other": "其他",
}

STATUS_LABELS = {
    "pending": "待确认",
    "todo": "待确认",
    "confirmed": "已确认",
    "done": "已处理",
    "rejected": "已驳回",
}

CJK_RE = re.compile(r"[\u4e00-\u9fff]")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Write E01 AI decisions back into review workbook and CSV deliverables.")
    parser.add_argument("--run-dir", required=True)
    parser.add_argument("--mode", default="review_only", choices=["review_only", "template_writeback"])
    parser.add_argument("--allow-validation-errors", action="store_true")
    return parser.parse_args()


def write_rows(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    ws.freeze_panes = "A2"
    for idx, header in enumerate(headers, 1):
        width = max(12, min(48, len(header) * 2 + 4))
        ws.column_dimensions[ws.cell(1, idx).column_letter].width = width


def build_review_rows(fields: list[dict[str, Any]], decisions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    lookup = {(f["sheet_name"], int(f["row_number"])): f for f in fields}
    rows: list[dict[str, Any]] = []
    for decision in decisions:
        key = (clean_text(decision.get("sheet_name")), int(decision.get("row_number")))
        field = lookup.get(key, {})
        rows.append({
            "工作表": key[0],
            "行号": key[1],
            "字段序号": field.get("field_order", ""),
            "字段分类": field.get("field_category", ""),
            "二级分类": field.get("field_subcategory", ""),
            "字段描述": field.get("field_desc", ""),
            "当前字段名": decision.get("current_field_name", field.get("field_name", "")),
            "推荐字段名": decision.get("recommended_field_name", ""),
            "数据类型": field.get("data_type", ""),
            "参考字段": field.get("reference_field", ""),
            "参考库表": field.get("reference_table", ""),
            "决策": decision_label(decision.get("decision", "")),
            "置信度": confidence_label(decision.get("confidence", "")),
            "标准化说明": localized_free_text(
                decision.get("reason", ""),
                f"根据字段描述、数据类型、参考表和参考字段检查，处理结果为“{decision_label(decision.get('decision', ''))}”。",
            ),
            "人工问题": localized_free_text(decision.get("human_question", ""), "需要人工确认该字段命名。"),
            "是否允许回写": yes_no_label(decision.get("writeback_allowed", False)),
        })
    return rows


def build_confirm_rows(review_items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in review_items:
        rows.append({
            "问题类型": issue_type_label(item.get("issue_type", "")),
            "严重级别": severity_label(item.get("severity", "")),
            "工作表": item.get("sheet_name", ""),
            "行号": item.get("row_number", ""),
            "问题描述": item.get("issue", ""),
            "标准化建议": item.get("suggestion", ""),
            "标准化说明": localized_free_text(item.get("ai_reason", ""), "该项需要人工确认后再回填。"),
            "用户意见": "",
            "处理状态": "待确认",
        })
    return rows


def build_table_review_rows(table_decisions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for decision in table_decisions:
        rows.append({
            "工作表": decision.get("sheet_name", ""),
            "当前表名": decision.get("current_table_name", ""),
            "推荐表名": decision.get("recommended_table_name", ""),
            "决策": decision_label(decision.get("decision", "")),
            "置信度": confidence_label(decision.get("confidence", "")),
            "标准化说明": localized_free_text(
                decision.get("reason", ""),
                f"根据目录页、实体标题和表名规范检查，处理结果为“{decision_label(decision.get('decision', ''))}”。",
            ),
            "人工问题": localized_free_text(decision.get("human_question", ""), "需要人工确认该表名。"),
        })
    return rows


def decision_label(value: Any) -> str:
    text = clean_text(value)
    return DECISION_LABELS.get(text, text)


def confidence_label(value: Any) -> str:
    text = clean_text(value)
    return CONFIDENCE_LABELS.get(text, text)


def severity_label(value: Any) -> str:
    text = clean_text(value)
    return SEVERITY_LABELS.get(text, text)


def issue_type_label(value: Any) -> str:
    text = clean_text(value)
    return ISSUE_TYPE_LABELS.get(text, text)


def gap_type_label(value: Any) -> str:
    text = clean_text(value)
    return GAP_TYPE_LABELS.get(text, text)


def status_label(value: Any) -> str:
    text = clean_text(value)
    return STATUS_LABELS.get(text, text)


def writeback_mode_label(value: str) -> str:
    return "模板回填" if value == "template_writeback" else "仅生成评审建议"


def yes_no_label(value: Any) -> str:
    return "是" if bool(value) else "否"


def localized_free_text(value: Any, fallback: str = "") -> str:
    text = clean_text(value)
    if not text:
        return ""
    if CJK_RE.search(text) and "??" not in text:
        return text
    return fallback or "说明内容需要改为中文。"


def build_gap_rows(gaps: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for gap in gaps:
        rows.append({
            "缺口类型": gap_type_label(gap.get("gap_type", "")),
            "来源工作表": gap.get("source_sheet", ""),
            "来源行号": gap.get("source_row", ""),
            "字段描述": gap.get("field_desc", ""),
            "推荐标准": gap.get("suggested_standard", ""),
            "标准化说明": localized_free_text(gap.get("reason", ""), "该标准缺口需要人工确认后沉淀到标准库。"),
            "状态": status_label(gap.get("status", "待确认")),
            "用户意见": "",
        })
    return rows


def normalized_header(value: Any) -> str:
    return clean_text(value).replace(" ", "")


def find_header_row(ws: Any, required_headers: list[str]) -> int | None:
    required = {normalized_header(h) for h in required_headers}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50)):
        values = {normalized_header(cell.value) for cell in row if clean_text(cell.value)}
        if required.issubset(values):
            return row[0].row
    return None


def header_map(ws: Any, row_number: int | None) -> dict[str, int]:
    if not row_number:
        return {}
    result: dict[str, int] = {}
    for cell in ws[row_number]:
        text = normalized_header(cell.value)
        if text:
            result[text] = cell.column
    return result


def should_apply_field_decision(decision: dict[str, Any], mode: str) -> bool:
    recommended = clean_text(decision.get("recommended_field_name"))
    decision_type = clean_text(decision.get("decision"))
    confidence = clean_text(decision.get("confidence"))
    writeback_allowed = bool(decision.get("writeback_allowed"))
    if not recommended or decision_type not in AUTO_FIELD_DECISIONS:
        return False
    if writeback_allowed and confidence in {"high", "medium"}:
        return True
    return mode == "template_writeback" and confidence == "high"


def should_apply_table_decision(decision: dict[str, Any], mode: str) -> bool:
    recommended = clean_text(decision.get("recommended_table_name"))
    decision_type = clean_text(decision.get("decision"))
    confidence = clean_text(decision.get("confidence"))
    if not recommended or decision_type not in AUTO_TABLE_DECISIONS:
        return False
    return mode == "template_writeback" and confidence == "high"


def not_applied_status(decision: dict[str, Any], recommended_key: str) -> str:
    decision_type = clean_text(decision.get("decision"))
    confidence = clean_text(decision.get("confidence"))
    recommended = clean_text(decision.get(recommended_key))
    if not recommended:
        return "未回填（无推荐值）"
    if decision_type in {"suggest_with_review", "need_human_confirm", "cannot_decide"}:
        return "未回填（待确认）"
    if confidence == "low":
        return "未回填（置信度低）"
    return "未回填"


def audit_record(
    object_type: str,
    status: str,
    sheet_name: str,
    cell_ref: str,
    row_number: Any,
    before: Any,
    after: Any,
    decision: Any,
    confidence: Any,
    reason: Any,
    human_question: Any = "",
) -> dict[str, Any]:
    return {
        "对象类型": object_type,
        "处理结果": status,
        "工作表": sheet_name,
        "单元格": cell_ref,
        "行号": row_number,
        "原值": clean_text(before),
        "标准化后": clean_text(after),
        "决策": decision_label(decision),
        "置信度": confidence_label(confidence),
        "说明": localized_free_text(reason, "根据字段、表名或目录标准化检查结果完成处理。"),
        "人工确认问题": localized_free_text(human_question, "需要人工确认后再处理。"),
    }


def fill_for_status(status: str) -> PatternFill:
    if status.startswith("已修改"):
        return CHANGED_FILL
    if status.startswith("已检查未改"):
        return KEPT_FILL
    return PENDING_FILL


def mark_changed_cell(cell: Any, old_value: Any, new_value: Any, reason: str) -> None:
    cell.fill = CHANGED_FILL
    note = [
        "标准化已修改",
        f"原值：{clean_text(old_value)}",
        f"新值：{clean_text(new_value)}",
    ]
    reason_text = localized_free_text(reason, "根据标准化检查结果修改。")
    if reason_text:
        note.append(f"说明：{reason_text}")
    cell.comment = Comment("\n".join(note), "Codex")


def write_audit_sheet(wb: Any, summary_rows: list[dict[str, Any]], audit_rows: list[dict[str, Any]]) -> None:
    if AUDIT_SHEET in wb.sheetnames:
        del wb[AUDIT_SHEET]
    ws = wb.create_sheet(AUDIT_SHEET, 0)
    ws["A1"] = "标准化结果总览"
    ws["A1"].font = Font(bold=True, size=14)

    ws.append([])
    ws.append(["指标", "值"])
    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    for row in summary_rows:
        ws.append([row.get("指标", ""), row.get("值", "")])

    detail_start = len(summary_rows) + 6
    ws.cell(detail_start, 1).value = "标准化处理明细"
    ws.cell(detail_start, 1).font = Font(bold=True, size=12)
    headers = ["对象类型", "处理结果", "工作表", "单元格", "行号", "原值", "标准化后", "决策", "置信度", "说明", "人工确认问题"]
    header_row = detail_start + 1
    for col, header in enumerate(headers, 1):
        cell = ws.cell(header_row, col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    for row_idx, row in enumerate(audit_rows, header_row + 1):
        status = clean_text(row.get("处理结果"))
        sheet_name = clean_text(row.get("工作表"))
        cell_ref = clean_text(row.get("单元格"))
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row_idx, col)
            cell.value = row.get(header, "")
            cell.fill = fill_for_status(status)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if header == "单元格" and sheet_name and cell_ref:
                cell.hyperlink = f"#{quote_cell_location(sheet_name, cell_ref)}"
                cell.font = HYPERLINK_FONT

    ws.freeze_panes = ws.cell(header_row + 1, 1).coordinate
    if audit_rows:
        ws.auto_filter.ref = f"A{header_row}:K{header_row + len(audit_rows)}"
    widths = {
        "A": 14,
        "B": 18,
        "C": 28,
        "D": 12,
        "E": 10,
        "F": 28,
        "G": 28,
        "H": 16,
        "I": 10,
        "J": 48,
        "K": 32,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def quote_sheet_location(sheet_name: str) -> str:
    escaped = sheet_name.replace("'", "''")
    return f"'{escaped}'!A1"


def quote_cell_location(sheet_name: str, cell_ref: str) -> str:
    escaped = sheet_name.replace("'", "''")
    return f"'{escaped}'!{cell_ref}"


def copy_row_style(ws: Any, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.border:
            target.border = copy(source.border)
        if source.fill:
            target.fill = copy(source.fill)
        if source.font:
            target.font = copy(source.font)


def replace_first_table_name(original: str, table_name: str) -> str:
    text = clean_text(original)
    if not text:
        return table_name
    match = re.match(r"^([^，,]*)(.*)$", text)
    suffix = match.group(2) if match else ""
    return f"{table_name}{suffix}"


def apply_template_writeback(
    wb: Any,
    fields: list[dict[str, Any]],
    decisions: list[dict[str, Any]],
    table_decisions: list[dict[str, Any]],
    entity_sheets: list[dict[str, Any]],
) -> dict[str, Any]:
    field_lookup = {(f["sheet_name"], int(f["row_number"])): f for f in fields}
    field_updates = 0
    skipped_field_decisions = 0
    audit_rows: list[dict[str, Any]] = []

    for decision in decisions:
        sheet_name = clean_text(decision.get("sheet_name"))
        row_number = int(decision.get("row_number"))
        field = field_lookup.get((sheet_name, row_number))
        if not field or sheet_name not in wb.sheetnames:
            skipped_field_decisions += 1
            audit_rows.append(audit_record(
                "字段名",
                "未回填（未找到位置）",
                sheet_name,
                "",
                row_number,
                decision.get("current_field_name", ""),
                decision.get("recommended_field_name", ""),
                decision.get("decision", ""),
                decision.get("confidence", ""),
                decision.get("reason", ""),
                decision.get("human_question", ""),
            ))
            continue
        ws = wb[sheet_name]
        cell_ref = field.get("source_cell_map", {}).get("field_name")
        if cell_ref:
            target = ws[cell_ref]
        else:
            header_row = find_header_row(ws, ["字段名", "字段描述"])
            field_col = header_map(ws, header_row).get(normalized_header("字段名"))
            if not field_col:
                skipped_field_decisions += 1
                audit_rows.append(audit_record(
                    "字段名",
                    "未回填（未找到字段名列）",
                    sheet_name,
                    "",
                    row_number,
                    decision.get("current_field_name", ""),
                    decision.get("recommended_field_name", ""),
                    decision.get("decision", ""),
                    decision.get("confidence", ""),
                    decision.get("reason", ""),
                    decision.get("human_question", ""),
                ))
                continue
            target = ws.cell(row_number, field_col)
        recommended = clean_text(decision.get("recommended_field_name"))
        before = clean_text(target.value)
        if not should_apply_field_decision(decision, "template_writeback"):
            skipped_field_decisions += 1
            audit_rows.append(audit_record(
                "字段名",
                not_applied_status(decision, "recommended_field_name"),
                sheet_name,
                target.coordinate,
                row_number,
                before,
                recommended,
                decision.get("decision", ""),
                decision.get("confidence", ""),
                decision.get("reason", ""),
                decision.get("human_question", ""),
            ))
            continue
        if before != recommended:
            target.value = recommended
            mark_changed_cell(target, before, recommended, decision.get("reason", ""))
            field_updates += 1
            status = "已修改"
        else:
            status = "已检查未改"
        audit_rows.append(audit_record(
            "字段名",
            status,
            sheet_name,
            target.coordinate,
            row_number,
            before,
            recommended,
            decision.get("decision", ""),
            decision.get("confidence", ""),
            decision.get("reason", ""),
            decision.get("human_question", ""),
        ))

    table_updates = 0
    catalog_updates = 0
    hyperlink_updates = 0
    appended_catalog_rows = 0
    table_by_sheet: dict[str, str] = {}

    table_decision_by_sheet: dict[str, dict[str, Any]] = {}
    for decision in table_decisions:
        sheet_name = clean_text(decision.get("sheet_name"))
        if sheet_name:
            table_decision_by_sheet[sheet_name] = decision
        if not should_apply_table_decision(decision, "template_writeback"):
            continue
        recommended = clean_text(decision.get("recommended_table_name"))
        if sheet_name and recommended:
            table_by_sheet[sheet_name] = recommended

    for sheet in entity_sheets:
        sheet_name = clean_text(sheet.get("sheet_name"))
        if sheet_name and sheet_name not in table_by_sheet:
            table_by_sheet[sheet_name] = clean_text(sheet.get("declared_table"))

    for sheet_name, table_name in table_by_sheet.items():
        if sheet_name not in wb.sheetnames or not table_name:
            continue
        decision = table_decision_by_sheet.get(sheet_name, {})
        applied_by_decision = should_apply_table_decision(decision, "template_writeback") if decision else False
        if applied_by_decision:
            ws = wb[sheet_name]
            before = clean_text(ws.cell(1, 1).value)
            new_title = replace_first_table_name(ws.cell(1, 1).value, table_name)
            if before != new_title:
                ws.cell(1, 1).value = new_title
                mark_changed_cell(ws.cell(1, 1), before, new_title, decision.get("reason", ""))
                table_updates += 1
                status = "已修改"
            else:
                status = "已检查未改"
            audit_rows.append(audit_record(
                "实体表名",
                status,
                sheet_name,
                "A1",
                1,
                before,
                new_title,
                decision.get("decision", ""),
                decision.get("confidence", ""),
                decision.get("reason", ""),
                decision.get("human_question", ""),
            ))

    catalog_sheet = "1.1 模型设计目录"
    if catalog_sheet in wb.sheetnames:
        ws = wb[catalog_sheet]
        header_row = find_header_row(ws, ["模型编号", "模型名称"])
        hmap = header_map(ws, header_row)
        model_col = hmap.get(normalized_header("模型名称"))
        table_col = hmap.get(normalized_header("实体表名(库名.表名)")) or hmap.get(normalized_header("实体表名"))
        if header_row and model_col:
            catalog_rows: dict[str, int] = {}
            last_data_row = header_row
            for row in range(header_row + 1, ws.max_row + 1):
                model_name = clean_text(ws.cell(row, model_col).value)
                if model_name:
                    catalog_rows[model_name] = row
                    last_data_row = row
            for sheet in entity_sheets:
                sheet_name = clean_text(sheet.get("sheet_name"))
                if not sheet_name or sheet_name not in wb.sheetnames:
                    continue
                table_name = table_by_sheet.get(sheet_name, "")
                row = catalog_rows.get(sheet_name)
                if not row:
                    row = last_data_row + 1
                    copy_row_style(ws, last_data_row, row, ws.max_column)
                    ws.cell(row, model_col).value = sheet_name
                    catalog_rows[sheet_name] = row
                    last_data_row = row
                    appended_catalog_rows += 1
                    audit_rows.append(audit_record(
                        "目录行",
                        "已修改（新增模型行）",
                        catalog_sheet,
                        ws.cell(row, model_col).coordinate,
                        row,
                        "",
                        sheet_name,
                        "",
                        "",
                        "目录页缺少该模型，已新增目录行。",
                    ))
                model_cell = ws.cell(row, model_col)
                target_location = quote_sheet_location(sheet_name)
                old_link = model_cell.hyperlink.location if model_cell.hyperlink else ""
                if not model_cell.hyperlink or model_cell.hyperlink.location != target_location:
                    model_cell.hyperlink = f"#{target_location}"
                    model_cell.style = "Hyperlink"
                    hyperlink_updates += 1
                    audit_rows.append(audit_record(
                        "目录跳转",
                        "已修改",
                        catalog_sheet,
                        model_cell.coordinate,
                        row,
                        old_link,
                        target_location,
                        "",
                        "",
                        "维护模型名称到实体工作表的跳转。",
                    ))
                if table_col and table_name and clean_text(ws.cell(row, table_col).value) != table_name:
                    cell = ws.cell(row, table_col)
                    before = clean_text(cell.value)
                    ws.cell(row, table_col).value = table_name
                    mark_changed_cell(cell, before, table_name, "维护目录页实体表名。")
                    catalog_updates += 1
                    audit_rows.append(audit_record(
                        "目录实体表名",
                        "已修改",
                        catalog_sheet,
                        cell.coordinate,
                        row,
                        before,
                        table_name,
                        table_decision_by_sheet.get(sheet_name, {}).get("decision", ""),
                        table_decision_by_sheet.get(sheet_name, {}).get("confidence", ""),
                        table_decision_by_sheet.get(sheet_name, {}).get("reason", "维护目录页实体表名。"),
                    ))

    return {
        "field_updates": field_updates,
        "skipped_field_decisions": skipped_field_decisions,
        "table_title_updates": table_updates,
        "catalog_table_updates": catalog_updates,
        "catalog_hyperlink_updates": hyperlink_updates,
        "catalog_rows_appended": appended_catalog_rows,
        "audit_rows": audit_rows,
    }


def main() -> None:
    args = parse_args()
    run_dir = to_project_path(args.run_dir)
    manifest = read_json(run_dir / "run_manifest.json")
    validation_path = run_dir / "04_validation" / "validation_report.json"
    if validation_path.exists():
        validation = read_json(validation_path)
        if validation.get("status") == "failed" and not args.allow_validation_errors:
            raise RuntimeError("validation_report.json is failed; rerun with --allow-validation-errors to write anyway.")

    model_path = to_project_path(manifest["inputs"]["model_design"])
    fields = read_jsonl(run_dir / "01_parse" / "entity_fields.jsonl")
    entity_sheets = read_jsonl(run_dir / "01_parse" / "entity_sheets.jsonl")
    decisions = read_jsonl(run_dir / "03_ai_decisions" / "ai_field_review_decisions.jsonl")
    table_decisions_path = run_dir / "03_ai_decisions" / "ai_table_review_decisions.jsonl"
    table_decisions = read_jsonl(table_decisions_path) if table_decisions_path.exists() else []
    review_items = read_jsonl(run_dir / "04_validation" / "review_items.jsonl")
    gaps = read_jsonl(run_dir / "03_ai_decisions" / "ai_standard_gaps.jsonl")

    if not decisions:
        raise RuntimeError("No AI field decisions found. Fill 03_ai_decisions/ai_field_review_decisions.jsonl first.")

    review_rows = build_review_rows(fields, decisions)
    table_review_rows = build_table_review_rows(table_decisions)
    confirm_rows = build_confirm_rows(review_items)
    gap_rows = build_gap_rows(gaps)

    if args.mode == "template_writeback":
        pending_path = run_dir / "05_writeback" / f"{model_path.stem}_待标准化.xlsx"
        final_path = run_dir / "06_deliverables" / f"{model_path.stem}_已标准化.xlsx"
        copy_file(model_path, pending_path)
        pending_hash = sha256_file(pending_path)
        source_hash = sha256_file(model_path)
        if pending_hash != source_hash:
            raise RuntimeError("待标准化副本与原始模型设计文件 hash 不一致，停止回填。")
        copy_file(pending_path, final_path)
    else:
        pending_path = run_dir / "05_writeback" / "workbook_draft.xlsx"
        final_path = run_dir / "06_deliverables" / "模型设计标准化建议版.xlsx"
        copy_file(model_path, pending_path)
        copy_file(model_path, final_path)

    wb = load_workbook(final_path)
    template_writeback_log: dict[str, Any] = {}
    audit_rows: list[dict[str, Any]] = []
    if args.mode == "template_writeback":
        template_writeback_log = apply_template_writeback(wb, fields, decisions, table_decisions, entity_sheets)
        audit_rows = template_writeback_log.pop("audit_rows", [])
    else:
        for sheet_name in REVIEW_SHEETS + LEGACY_REVIEW_SHEETS:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]

    summary_rows = [
        {"指标": "模型设计文件", "值": model_path.name},
        {"指标": "字段决策数", "值": len(decisions)},
        {"指标": "表名决策数", "值": len(table_decisions)},
        {"指标": "待人工确认数", "值": len(confirm_rows)},
        {"指标": "待沉淀标准数", "值": len(gap_rows)},
        {"指标": "输出模式", "值": writeback_mode_label(args.mode)},
        {"指标": "字段原位更新数", "值": template_writeback_log.get("field_updates", 0)},
        {"指标": "表名标题更新数", "值": template_writeback_log.get("table_title_updates", 0)},
        {"指标": "目录表名更新数", "值": template_writeback_log.get("catalog_table_updates", 0)},
        {"指标": "结果清单", "值": "已生成“0 标准化结果”工作表" if args.mode == "template_writeback" else "已生成评审工作表"},
        {"指标": "生成时间", "值": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
    ]
    if args.mode == "review_only":
        write_rows(
            wb.create_sheet("字段标准化建议"),
            ["工作表", "行号", "字段序号", "字段分类", "二级分类", "字段描述", "当前字段名", "推荐字段名", "数据类型", "参考字段", "参考库表", "决策", "置信度", "标准化说明", "人工问题", "是否允许回写"],
            review_rows,
        )
        write_rows(
            wb.create_sheet("表名标准化建议"),
            ["工作表", "当前表名", "推荐表名", "决策", "置信度", "标准化说明", "人工问题"],
            table_review_rows,
        )
        write_rows(
            wb.create_sheet("待人工确认"),
            ["问题类型", "严重级别", "工作表", "行号", "问题描述", "标准化建议", "标准化说明", "用户意见", "处理状态"],
            confirm_rows,
        )
        write_rows(
            wb.create_sheet("待沉淀标准"),
            ["缺口类型", "来源工作表", "来源行号", "字段描述", "推荐标准", "标准化说明", "状态", "用户意见"],
            gap_rows,
        )
        write_rows(wb.create_sheet("运行摘要"), ["指标", "值"], summary_rows)
    else:
        write_audit_sheet(wb, summary_rows, audit_rows)
    wb.save(final_path)

    write_csv(
        run_dir / "06_deliverables" / "待人工确认清单.csv",
        confirm_rows,
        ["问题类型", "严重级别", "工作表", "行号", "问题描述", "标准化建议", "标准化说明", "用户意见", "处理状态"],
    )
    write_csv(
        run_dir / "06_deliverables" / "待沉淀标准清单.csv",
        gap_rows,
        ["缺口类型", "来源工作表", "来源行号", "字段描述", "推荐标准", "标准化说明", "状态", "用户意见"],
    )

    plan = {
        "mode": args.mode,
        "actions": [
            {"action": "copy_source_workbook", "target": rel(pending_path)},
            {"action": "create_pending_copy", "target": rel(pending_path), "hash_matches_source": True} if args.mode == "template_writeback" else {"action": "create_draft_copy", "target": rel(pending_path)},
            {"action": "rename_to_standardized", "target": rel(final_path)} if args.mode == "template_writeback" else {"action": "copy_to_review_workbook", "target": rel(final_path)},
            {"action": "template_writeback", "details": template_writeback_log} if args.mode == "template_writeback" else {"action": "review_only", "details": "original field names kept unchanged"},
            {"action": "add_excel_audit_sheet", "details": "0 标准化结果"} if args.mode == "template_writeback" else {"action": "replace_review_sheets", "sheets": REVIEW_SHEETS},
            {"action": "export_csv", "files": ["待人工确认清单.csv", "待沉淀标准清单.csv"]},
        ],
        "overwrite_original_field_names": args.mode == "template_writeback",
    }
    log = {
        "status": "completed",
        "mode": args.mode,
        "review_rows": len(review_rows),
        "table_review_rows": len(table_review_rows),
        "confirm_rows": len(confirm_rows),
        "gap_rows": len(gap_rows),
        **template_writeback_log,
        "pending_workbook": rel(pending_path),
        "final_workbook": rel(final_path),
    }
    write_json(run_dir / "05_writeback" / "writeback_plan.json", plan)
    write_json(run_dir / "05_writeback" / "writeback_log.json", log)

    manifest["status"] = "writeback_completed"
    if args.mode == "template_writeback":
        manifest["outputs"]["template_writeback_excel"] = rel(final_path)
    else:
        manifest["outputs"]["review_excel"] = rel(final_path)
    manifest["outputs"]["confirm_csv"] = rel(run_dir / "06_deliverables" / "待人工确认清单.csv")
    manifest["outputs"]["standard_gap_csv"] = rel(run_dir / "06_deliverables" / "待沉淀标准清单.csv")
    write_json(run_dir / "run_manifest.json", manifest)

    print(f"[OK] output_excel={rel(final_path)}")
    print(f"[OK] confirm_rows={len(confirm_rows)} standard_gap_rows={len(gap_rows)}")


if __name__ == "__main__":
    main()
