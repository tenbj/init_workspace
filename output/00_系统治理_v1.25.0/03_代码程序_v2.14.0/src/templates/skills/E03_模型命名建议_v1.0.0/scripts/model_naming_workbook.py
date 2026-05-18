#!/usr/bin/env python3
from __future__ import annotations

import argparse
import copy
import json
import shutil
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - runtime environment guard
    print("Missing dependency: openpyxl. Install it before using this script.", file=sys.stderr)
    raise SystemExit(2) from exc


FIELD_NAME_HEADERS = {"字段名", "字段英文名", "英文名", "字段编码", "字段code", "字段代码"}
RECOMMEND_HEADER = "推荐字段名"
REMARK_HEADERS = {"备注", "说明", "字段说明", "描述"}
HEADER_SCAN_ROWS = 50
MAX_SCAN_COLS = 80


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def compact(value: Any) -> str:
    return "".join(text(value).split()).lower()


def truncate(value: Any, limit: int = 240) -> str:
    raw = text(value)
    return raw if len(raw) <= limit else raw[:limit] + "..."


def ensure_xlsx(path: Path) -> None:
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"Only .xlsx is supported: {path}")
    if not path.exists():
        raise FileNotFoundError(path)


def is_field_name_header(value: Any) -> bool:
    normalized = compact(value)
    if normalized == compact(RECOMMEND_HEADER):
        return False
    return normalized in {compact(item) for item in FIELD_NAME_HEADERS}


def is_remark_header(value: Any) -> bool:
    return compact(value) in {compact(item) for item in REMARK_HEADERS}


def find_header_row_and_field_col(ws: Any) -> tuple[int, int] | None:
    max_row = min(ws.max_row or 0, HEADER_SCAN_ROWS)
    max_col = min(ws.max_column or 0, MAX_SCAN_COLS)
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            if is_field_name_header(ws.cell(row, col).value):
                return row, col
    return None


def header_cells(ws: Any, header_row: int) -> list[dict[str, Any]]:
    headers: list[dict[str, Any]] = []
    for col in range(1, min(ws.max_column or 0, MAX_SCAN_COLS) + 1):
        value = text(ws.cell(header_row, col).value)
        if value:
            headers.append({"col": col, "header": value})
    return headers


def find_header_col(ws: Any, header_row: int, predicate: Any) -> int | None:
    for col in range(1, ws.max_column + 1):
        if predicate(ws.cell(header_row, col).value):
            return col
    return None


def row_has_values(ws: Any, row: int, headers: list[dict[str, Any]]) -> bool:
    return any(text(ws.cell(row, item["col"]).value) for item in headers)


def field_rows(ws: Any, header_row: int, field_col: int, headers: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    blank_streak = 0
    for row in range(header_row + 1, ws.max_row + 1):
        if not row_has_values(ws, row, headers):
            blank_streak += 1
            if blank_streak >= 20:
                break
            continue
        blank_streak = 0
        current_name = text(ws.cell(row, field_col).value)
        values = {
            item["header"]: truncate(ws.cell(row, item["col"]).value)
            for item in headers
            if text(ws.cell(row, item["col"]).value)
        }
        rows.append({"row": row, "current_field_name": current_name, "values": values})
    return rows


def detect_tabular_header_row(ws: Any) -> int | None:
    max_row = min(ws.max_row or 0, 20)
    max_col = min(ws.max_column or 0, MAX_SCAN_COLS)
    best_row = None
    best_count = 0
    for row in range(1, max_row + 1):
        count = sum(1 for col in range(1, max_col + 1) if text(ws.cell(row, col).value))
        if count > best_count:
            best_row = row
            best_count = count
    return best_row if best_count >= 2 else None


def extract_standard_sheet(ws: Any, max_rows: int) -> dict[str, Any] | None:
    header_row = detect_tabular_header_row(ws)
    if not header_row:
        return None
    headers = header_cells(ws, header_row)
    records: list[dict[str, Any]] = []
    for row in range(header_row + 1, min(ws.max_row, header_row + max_rows) + 1):
        record = {
            item["header"]: truncate(ws.cell(row, item["col"]).value)
            for item in headers
            if text(ws.cell(row, item["col"]).value)
        }
        if record:
            records.append({"row": row, "values": record})
    return {"sheet": ws.title, "header_row": header_row, "headers": headers, "records": records}


def extract_context(args: argparse.Namespace) -> int:
    standard_path = Path(args.standard_lib).resolve()
    model_path = Path(args.model).resolve()
    out_path = Path(args.out).resolve()
    ensure_xlsx(standard_path)
    ensure_xlsx(model_path)

    model_wb = load_workbook(model_path, data_only=True)
    standard_wb = load_workbook(standard_path, data_only=True)

    model_sheets: list[dict[str, Any]] = []
    for ws in model_wb.worksheets:
        located = find_header_row_and_field_col(ws)
        if not located:
            continue
        header_row, field_col = located
        headers = header_cells(ws, header_row)
        model_sheets.append(
            {
                "sheet": ws.title,
                "table_name_a1": truncate(ws["A1"].value),
                "header_row": header_row,
                "field_name_col": field_col,
                "headers": headers,
                "fields": field_rows(ws, header_row, field_col, headers),
            }
        )

    standard_sheets: list[dict[str, Any]] = []
    for ws in standard_wb.worksheets:
        extracted = extract_standard_sheet(ws, args.standard_max_rows)
        if extracted:
            standard_sheets.append(extracted)

    payload = {
        "model_file": str(model_path),
        "standard_lib_file": str(standard_path),
        "model_sheets": model_sheets,
        "standard_sheets": standard_sheets,
    }
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"context={out_path}")
    print(f"model_sheets={len(model_sheets)}")
    print(f"standard_sheets={len(standard_sheets)}")
    return 0


def normalize_recommendations(data: dict[str, Any]) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for sheet_item in data.get("sheets", []):
        sheet = text(sheet_item.get("sheet"))
        if not sheet:
            continue
        result.setdefault(sheet, {"table": "", "fields": []})
        result[sheet]["table"] = text(
            sheet_item.get("recommended_table_name")
            or sheet_item.get("table_recommendation")
            or sheet_item.get("recommended_name")
        )
        result[sheet]["fields"].extend(sheet_item.get("fields", []))

    for table in data.get("tables", []):
        sheet = text(table.get("sheet"))
        if sheet:
            result.setdefault(sheet, {"table": "", "fields": []})
            result[sheet]["table"] = text(table.get("recommended_table_name") or table.get("recommended_name"))

    for field in data.get("fields", []):
        sheet = text(field.get("sheet"))
        if sheet:
            result.setdefault(sheet, {"table": "", "fields": []})
            result[sheet]["fields"].append(field)
    return result


def copy_cell_style(source: Any, target: Any) -> None:
    if source.has_style:
        target._style = copy.copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy.copy(source.alignment)


def ensure_recommend_col(ws: Any, header_row: int, field_col: int) -> tuple[int, bool]:
    existing = find_header_col(ws, header_row, lambda value: compact(value) == compact(RECOMMEND_HEADER))
    if existing:
        return existing, False

    insert_at = field_col + 1
    ws.insert_cols(insert_at)
    for row in range(1, ws.max_row + 1):
        copy_cell_style(ws.cell(row, insert_at - 1), ws.cell(row, insert_at))
    ws.cell(header_row, insert_at).value = RECOMMEND_HEADER
    left_letter = get_column_letter(insert_at - 1)
    new_letter = get_column_letter(insert_at)
    ws.column_dimensions[new_letter].width = ws.column_dimensions[left_letter].width
    return insert_at, True


def ensure_remark_col(ws: Any, header_row: int) -> int:
    existing = find_header_col(ws, header_row, is_remark_header)
    if existing:
        return existing
    new_col = ws.max_column + 1
    ws.cell(header_row, new_col).value = "备注"
    copy_cell_style(ws.cell(header_row, new_col - 1), ws.cell(header_row, new_col))
    return new_col


def recommendation_for_row(
    row: int,
    current_name: str,
    by_row: dict[int, str],
    by_name: dict[str, str],
) -> str:
    if row in by_row:
        return by_row[row]
    return by_name.get(compact(current_name), "")


def append_note(existing: Any, note: str) -> str:
    current = text(existing)
    if not current:
        return note
    if note in current:
        return current
    return current + "\n" + note


def write_recommendations(args: argparse.Namespace) -> int:
    model_path = Path(args.model).resolve()
    recommendations_path = Path(args.recommendations).resolve()
    output_path = Path(args.output).resolve() if args.output else model_path.with_name(model_path.stem + "_命名建议.xlsx")
    ensure_xlsx(model_path)
    if not recommendations_path.exists():
        raise FileNotFoundError(recommendations_path)
    if output_path.exists() and not args.overwrite:
        raise FileExistsError(f"Output already exists: {output_path}. Use --overwrite or choose another --output.")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(model_path, output_path)

    data = json.loads(recommendations_path.read_text(encoding="utf-8"))
    recommendations = normalize_recommendations(data)
    wb = load_workbook(output_path)
    touched_sheets = 0
    touched_fields = 0

    for ws in wb.worksheets:
        if ws.title not in recommendations:
            continue
        located = find_header_row_and_field_col(ws)
        if not located:
            raise ValueError(f"Cannot find field name header in sheet: {ws.title}")
        header_row, field_col = located
        headers = header_cells(ws, header_row)
        rows = field_rows(ws, header_row, field_col, headers)
        recommend_col, _ = ensure_recommend_col(ws, header_row, field_col)
        remark_col = ensure_remark_col(ws, header_row)

        sheet_recs = recommendations[ws.title]
        by_row: dict[int, str] = {}
        by_name: dict[str, str] = {}
        for item in sheet_recs.get("fields", []):
            suggested = text(
                item.get("recommended_field_name")
                or item.get("recommendation")
                or item.get("recommended_name")
            )
            if not suggested:
                continue
            if item.get("row") is not None:
                by_row[int(item["row"])] = suggested
            current = text(item.get("current_field_name") or item.get("field_name"))
            if current:
                by_name[compact(current)] = suggested

        for item in rows:
            suggested = recommendation_for_row(item["row"], item["current_field_name"], by_row, by_name)
            if suggested:
                ws.cell(item["row"], recommend_col).value = suggested
                touched_fields += 1

        table_name = text(sheet_recs.get("table"))
        if table_name and rows:
            note = f"推荐表名：{table_name}"
            first_row = rows[0]["row"]
            ws.cell(first_row, remark_col).value = append_note(ws.cell(first_row, remark_col).value, note)

        touched_sheets += 1

    wb.save(output_path)
    print(f"output={output_path}")
    print(f"touched_sheets={touched_sheets}")
    print(f"touched_fields={touched_fields}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Extract model naming context and write naming recommendations to a copied workbook.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    extract = subparsers.add_parser("extract", help="Extract compact JSON context from standard library and model workbook.")
    extract.add_argument("--standard-lib", required=True, help="Path to standard library .xlsx")
    extract.add_argument("--model", required=True, help="Path to model design .xlsx")
    extract.add_argument("--out", required=True, help="Output context JSON path")
    extract.add_argument("--standard-max-rows", type=int, default=2000, help="Maximum rows per standard sheet")
    extract.set_defaults(func=extract_context)

    write = subparsers.add_parser("write", help="Copy model workbook and write recommendation JSON into it.")
    write.add_argument("--model", required=True, help="Path to model design .xlsx")
    write.add_argument("--recommendations", required=True, help="AI recommendations JSON path")
    write.add_argument("--output", help="Output .xlsx path; defaults to *_命名建议.xlsx next to model")
    write.add_argument("--overwrite", action="store_true", help="Allow overwriting output path")
    write.set_defaults(func=write_recommendations)
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    try:
        return args.func(args)
    except Exception as exc:
        print(f"error={exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
